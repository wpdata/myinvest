"""
MyInvest V0.3 - Excel Report Generator (T028)
Trading records export with conditional formatting using openpyxl.
"""

import logging
from datetime import datetime
from typing import Dict, Any, List
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """Excel report generator with 3 sheets and conditional formatting.

    Sheets:
    1. Trade Details - All trades with profit/loss
    2. Monthly Summary - Aggregated monthly statistics
    3. Position Analysis - Position holding periods and P&L
    """

    def __init__(self):
        """Initialize Excel report generator."""
        pass

    def generate_trading_records(
        self,
        backtest_result: Dict[str, Any],
        output_path: str
    ) -> str:
        """Generate Excel trading records report.

        Args:
            backtest_result: Backtest result dict
            output_path: Output Excel file path

        Returns:
            str: Path to generated Excel file
        """
        logger.info(f"[ExcelGenerator] Generating report: {output_path}")

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Sheet 1: Trade Details
        self._create_trade_details_sheet(wb, backtest_result)

        # Sheet 2: Monthly Summary
        self._create_monthly_summary_sheet(wb, backtest_result)

        # Sheet 3: Position Analysis
        self._create_position_analysis_sheet(wb, backtest_result)

        # Save workbook
        wb.save(output_path)

        logger.info(f"[ExcelGenerator] Report generated: {output_path}")
        return output_path

    def _create_trade_details_sheet(
        self,
        wb: Workbook,
        result: Dict[str, Any]
    ):
        """Create Sheet 1: Trade Details."""
        ws = wb.create_sheet("交易明细")

        # Headers
        headers = ['日期', '股票代码', '操作', '价格', '数量', '总金额', '盈亏', '累计收益']
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Add trade data
        trade_log = result.get('trade_log', [])

        for trade in trade_log:
            row = [
                trade.get('timestamp', ''),
                trade.get('symbol', ''),
                '买入' if trade.get('action') == 'BUY' else '卖出',
                trade.get('price', 0),
                trade.get('quantity', 0),
                trade.get('total_value', 0),
                trade.get('profit_loss', 0),
                trade.get('cumulative_pnl', 0)
            ]
            ws.append(row)

        # Apply conditional formatting to profit/loss column (G)
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=7)  # Column G (盈亏)

            if cell.value and cell.value > 0:
                cell.font = Font(color='00B050')  # Green
            elif cell.value and cell.value < 0:
                cell.font = Font(color='FF0000')  # Red

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

    def _create_monthly_summary_sheet(
        self,
        wb: Workbook,
        result: Dict[str, Any]
    ):
        """Create Sheet 2: Monthly Summary."""
        ws = wb.create_sheet("月度汇总")

        # Headers
        headers = ['月份', '总收益', '收益率', '交易次数', '胜率']
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Group trades by month
        trade_log = result.get('trade_log', [])
        monthly_data = self._aggregate_by_month(trade_log)

        for month, stats in monthly_data.items():
            row = [
                month,
                stats['total_pnl'],
                stats['return_pct'],
                stats['trade_count'],
                stats['win_rate']
            ]
            ws.append(row)

        # Format numbers
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=2).number_format = '¥#,##0.00'  # 总收益
            ws.cell(row=row_idx, column=3).number_format = '0.00%'       # 收益率
            ws.cell(row=row_idx, column=5).number_format = '0.00%'       # 胜率

    def _create_position_analysis_sheet(
        self,
        wb: Workbook,
        result: Dict[str, Any]
    ):
        """Create Sheet 3: Position Analysis."""
        ws = wb.create_sheet("持仓分析")

        # Headers
        headers = ['股票代码', '持有天数', '平均成本', '卖出价格', '盈亏', '收益率']
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        header_font = Font(color='000000', bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Analyze positions
        positions = self._analyze_positions(result.get('trade_log', []))

        for pos in positions:
            row = [
                pos['symbol'],
                pos['holding_days'],
                pos['avg_cost'],
                pos['sell_price'],
                pos['pnl'],
                pos['return_pct']
            ]
            ws.append(row)

            # Conditional formatting for P&L
            row_idx = ws.max_row
            pnl_cell = ws.cell(row=row_idx, column=5)
            if pnl_cell.value > 0:
                pnl_cell.font = Font(color='00B050', bold=True)
            elif pnl_cell.value < 0:
                pnl_cell.font = Font(color='FF0000', bold=True)

        # Format numbers
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=3).number_format = '¥#,##0.00'
            ws.cell(row=row_idx, column=4).number_format = '¥#,##0.00'
            ws.cell(row=row_idx, column=5).number_format = '¥#,##0.00'
            ws.cell(row=row_idx, column=6).number_format = '0.00%'

    def _aggregate_by_month(self, trade_log: List[Dict]) -> Dict:
        """Aggregate trades by month."""
        monthly = {}

        for trade in trade_log:
            timestamp = trade.get('timestamp', '')
            if timestamp:
                month = timestamp[:7]  # YYYY-MM

                if month not in monthly:
                    monthly[month] = {
                        'total_pnl': 0,
                        'return_pct': 0,
                        'trade_count': 0,
                        'wins': 0,
                        'win_rate': 0
                    }

                monthly[month]['trade_count'] += 1
                pnl = trade.get('profit_loss', 0)
                monthly[month]['total_pnl'] += pnl

                if pnl > 0:
                    monthly[month]['wins'] += 1

        # Calculate win rates
        for month in monthly:
            total = monthly[month]['trade_count']
            wins = monthly[month]['wins']
            monthly[month]['win_rate'] = (wins / total) if total > 0 else 0

        return monthly

    def _analyze_positions(self, trade_log: List[Dict]) -> List[Dict]:
        """Analyze position holding periods and P&L."""
        positions = []
        open_positions = {}

        for trade in trade_log:
            symbol = trade.get('symbol')
            action = trade.get('action')

            if action == 'BUY':
                if symbol not in open_positions:
                    open_positions[symbol] = {
                        'entry_date': trade.get('timestamp'),
                        'entry_price': trade.get('price'),
                        'quantity': trade.get('quantity', 0)
                    }
            elif action == 'SELL' and symbol in open_positions:
                pos = open_positions[symbol]
                entry_date = pos['entry_date']
                exit_date = trade.get('timestamp')

                # Calculate holding days (simplified)
                holding_days = 1  # Placeholder

                pnl = trade.get('profit_loss', 0)
                return_pct = pnl / (pos['entry_price'] * pos['quantity']) if pos['entry_price'] > 0 else 0

                positions.append({
                    'symbol': symbol,
                    'holding_days': holding_days,
                    'avg_cost': pos['entry_price'],
                    'sell_price': trade.get('price'),
                    'pnl': pnl,
                    'return_pct': return_pct
                })

                del open_positions[symbol]

        return positions
